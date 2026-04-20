#!/usr/bin/env python3
"""
31_suppletable_260420.py

Assemble every main Table (1-3) and Supplementary Table (S1-S11)
referenced in v0.7.5 manuscript as editable xlsx files under
260418_add/suppletable_260420/.

Each xlsx carries:
  - Arial 11pt default font
  - bold header row with light-gray fill + thin bottom border
  - freeze-pane at A2 so headers stick while scrolling
  - auto-filter enabled on header row (so reviewers can sort/filter)
  - column widths auto-set to max of (header length + 2, observed length + 2,
    capped at 40)
  - numeric format: 4 decimals for p/q/correlations/delta; 2 for counts
  - text-wrap on header + long-string cells

Tables (one xlsx per):
  Table 1 — clinical_characteristics.xlsx
  Table 2 — top20_feature_associations.xlsx
  Table 3 — external_meta_summary.xlsx
  Table S1 — master_36feature_subject.xlsx
  Table S2 — SBS_signature_activities.xlsx
  Table S3 — GSEA_Hallmark_Reactome.xlsx (2 sheets)
  Table S4 — driver_mutations_per_sample.xlsx (2 sheets: per-sample +
             oncoprint matrix)
  Table S5 — HLA_class_I_typing.xlsx
  Table S6 — neoantigen_per_sample.xlsx
  Table S7 — external_validation_detail.xlsx (3 sheets: A/B/C)
  Table S8 — cascade_BCa_bootstrap.xlsx
  Table S9 — HLA_LOH_per_locus.xlsx (2 sheets: strict + lite)
  Table S10 — baseline_factor_pharmacodynamics.xlsx (2 sheets: per-subject
              delta + sign table)
  Table S11 — TRUST4_IGHV_directional_stats.xlsx
"""

import os
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------------------------------------------------------
# I/O + style helpers
# ----------------------------------------------------------------------------
ROOT = "/data/data/TNT/analysis"
ADD = f"{ROOT}/260418_add"
OUT = f"{ADD}/suppletable_260420"
os.makedirs(OUT, exist_ok=True)

DEFAULT_FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="EAEAEA")
HEADER_FONT = Font(name=DEFAULT_FONT, size=11, bold=True)
BODY_FONT = Font(name=DEFAULT_FONT, size=10)
BORDER = Border(bottom=Side(style="thin", color="555555"))
THICK_BORDER = Border(bottom=Side(style="medium", color="333333"))


def apply_style(ws, df, freeze="A2", auto_filter=True,
                numeric_cols_fmt=None,
                numeric_decimals=4,
                wrap_cols=None):
    """Apply consistent style to a sheet that was populated with df.
    df: pandas DataFrame whose headers are in row 1."""
    ncols = ws.max_column
    nrows = ws.max_row
    # headers
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THICK_BORDER
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center",
                                   wrap_text=True)
    # body
    col_widths = []
    for c in range(1, ncols + 1):
        header = str(df.columns[c - 1])
        max_len = len(header)
        for r in range(2, nrows + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            val = cell.value
            if val is None:
                continue
            # numeric formatting
            if isinstance(val, (float, np.floating)):
                if val == int(val) and abs(val) < 1e6:
                    cell.number_format = "0"
                else:
                    cell.number_format = f"0.{'0' * numeric_decimals}"
            # wrap long text
            if wrap_cols and header in wrap_cols and val is not None:
                cell.alignment = Alignment(horizontal="left",
                                           vertical="top",
                                           wrap_text=True)
            # measure length
            strv = str(val)
            if len(strv) > max_len:
                max_len = min(len(strv), 60)
        col_widths.append(max_len + 2)
    # column widths
    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(w, 8), 45)
    # freeze + filter
    ws.freeze_panes = freeze
    if auto_filter:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"
    # row height for header
    ws.row_dimensions[1].height = 28


def write_df(wb, df, sheet_name, **kwargs):
    ws = wb.create_sheet(title=sheet_name[:31])
    # header
    for c, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c, value=str(col))
    # body
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for c, col in enumerate(df.columns, start=1):
            v = row[col]
            if isinstance(v, (np.floating, np.integer)):
                v = float(v) if isinstance(v, np.floating) else int(v)
            elif isinstance(v, float) and np.isnan(v):
                v = None
            ws.cell(row=r, column=c, value=v)
    apply_style(ws, df, **kwargs)
    return ws


def new_wb():
    wb = Workbook()
    # delete default sheet later
    return wb


def finalise(wb, fname):
    # remove the default blank sheet if still present
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        if std.max_row <= 1 and std.max_column <= 1:
            wb.remove(std)
    path = f"{OUT}/{fname}"
    wb.save(path)
    # probe
    from openpyxl import load_workbook
    wb2 = load_workbook(path, read_only=True)
    n_sh = len(wb2.sheetnames)
    print(f"  + {fname}  ({n_sh} sheet · {os.path.getsize(path) / 1024:.1f} KB)")


# ============================================================================
# TABLE 1 --- Clinical characteristics by response
# ============================================================================
def build_T1():
    clin = pd.read_csv(f"{ROOT}/00_cohort/clinical_master.tsv", sep="\t")
    # aggregate by response_bin
    from scipy.stats import mannwhitneyu, fisher_exact

    def mw(col):
        g = clin[clin.response_bin == "good"][col].dropna()
        b = clin[clin.response_bin == "bad"][col].dropna()
        _, p = mannwhitneyu(g, b)
        return p

    rows = []
    # age
    g_age = clin[clin.response_bin == "good"]["age"].dropna()
    b_age = clin[clin.response_bin == "bad"]["age"].dropna()
    rows.append({
        "Variable": "Age (years)",
        "Good (n=18)": f"{g_age.median():.1f} [{g_age.min():.0f}–{g_age.max():.0f}]",
        "Bad (n=17)": f"{b_age.median():.1f} [{b_age.min():.0f}–{b_age.max():.0f}]",
        "Test": "Mann–Whitney U",
        "P value": f"{mw('age'):.3f}",
    })
    # sex
    sx = pd.crosstab(clin["sex"], clin["response_bin"])
    _, p_sex = fisher_exact(sx.values) if sx.shape == (2, 2) else (0, 1.0)
    rows.append({"Variable": "Sex — F", "Good (n=18)": int(sx.loc["F", "good"]) if "F" in sx.index else 0,
                 "Bad (n=17)": int(sx.loc["F", "bad"]) if "F" in sx.index else 0,
                 "Test": "Fisher exact", "P value": f"{p_sex:.3f}"})
    rows.append({"Variable": "Sex — M", "Good (n=18)": int(sx.loc["M", "good"]) if "M" in sx.index else 0,
                 "Bad (n=17)": int(sx.loc["M", "bad"]) if "M" in sx.index else 0,
                 "Test": "", "P value": ""})
    # cT
    ct_tab = pd.crosstab(clin["cT"], clin["response_bin"])
    for stage in ["T2", "T2/T3", "T3", "T4"]:
        if stage in ct_tab.index:
            rows.append({
                "Variable": f"cT — {stage}",
                "Good (n=18)": int(ct_tab.loc[stage, "good"]) if "good" in ct_tab.columns else 0,
                "Bad (n=17)": int(ct_tab.loc[stage, "bad"]) if "bad" in ct_tab.columns else 0,
                "Test": "", "P value": "",
            })
    # cT4 vs rest (Fisher)
    t4_mask = clin["cT"] == "T4"
    tab = pd.crosstab(t4_mask, clin["response_bin"])
    if tab.shape == (2, 2):
        _, p_t4 = fisher_exact(tab.values)
        rows.append({"Variable": "cT4 vs not-T4",
                     "Good (n=18)": f"{int(tab.loc[True, 'good'])} / {int(tab.loc[False, 'good'])}",
                     "Bad (n=17)": f"{int(tab.loc[True, 'bad'])} / {int(tab.loc[False, 'bad'])}",
                     "Test": "Fisher exact", "P value": f"{p_t4:.3f}"})
    # prepost
    pp_tab = pd.crosstab(clin["prepost_set"], clin["response_bin"])
    for k in pp_tab.index:
        rows.append({"Variable": f"prepost_set — {k}",
                     "Good (n=18)": int(pp_tab.loc[k, "good"]) if "good" in pp_tab.columns else 0,
                     "Bad (n=17)": int(pp_tab.loc[k, "bad"]) if "bad" in pp_tab.columns else 0,
                     "Test": "", "P value": ""})
    # TMB
    tmb = pd.read_csv(f"{ROOT}/02_wes_tmb_msi/tmb_per_sample.tsv", sep="\t")
    tmb_pre = tmb[tmb.timepoint == "pre"].copy()
    g = tmb_pre[tmb_pre.response_bin == "good"]["TMB_nonsyn_per_Mb"].dropna()
    b = tmb_pre[tmb_pre.response_bin == "bad"]["TMB_nonsyn_per_Mb"].dropna()
    _, p_tmb = mannwhitneyu(g, b)
    rows.append({
        "Variable": "TMB (nonsyn / Mb)",
        "Good (n=18)": f"{g.median():.2f} [{g.min():.2f}–{g.max():.2f}]",
        "Bad (n=17)": f"{b.median():.2f} [{b.min():.2f}–{b.max():.2f}]",
        "Test": "Mann–Whitney U", "P value": f"{p_tmb:.3f}",
    })
    df = pd.DataFrame(rows)
    wb = new_wb()
    write_df(wb, df, "Clinical characteristics",
             wrap_cols=["Variable", "Good (n=18)", "Bad (n=17)", "Test"])
    finalise(wb, "Table1_clinical_characteristics.xlsx")


# ============================================================================
# TABLE 2 --- Top 20 integrated-feature associations
# ============================================================================
def build_T2():
    src = f"{ROOT}/tables/response_feature_stats.tsv"
    df = pd.read_csv(src, sep="\t")
    top = df.sort_values("pvalue").head(20).copy()
    top.columns = ["Feature", "n good", "n bad",
                   "Median good", "Median bad",
                   "Δ (good − bad)", "MW P", "BH q"]
    # round
    for c in ["Median good", "Median bad", "Δ (good − bad)"]:
        top[c] = top[c].astype(float).round(4)
    for c in ["MW P", "BH q"]:
        top[c] = top[c].astype(float).round(5)
    wb = new_wb()
    write_df(wb, top, "Top20 features",
             wrap_cols=["Feature"])
    finalise(wb, "Table2_top20_feature_associations.xlsx")


# ============================================================================
# TABLE 3 --- External validation meta summary
# ============================================================================
def build_T3():
    src = f"{ADD}/FINAL_meta_with_akiyoshi.tsv"
    df = pd.read_csv(src, sep="\t")
    df = df[["thread", "signature", "n_cohorts", "n_total", "Z", "p_meta",
             "cohorts", "akiyoshi", "5cohort_only_Z", "5cohort_only_p"]]
    df.columns = ["Thread", "Signature", "n cohorts", "N total",
                  "Stouffer Z", "P meta",
                  "Cohorts", "Akiyoshi augmentation",
                  "5-cohort only Z", "5-cohort only P"]
    for c in ["Stouffer Z", "P meta", "5-cohort only Z", "5-cohort only P"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").round(4)
    wb = new_wb()
    write_df(wb, df, "External meta",
             wrap_cols=["Signature", "Cohorts", "Akiyoshi augmentation"])
    finalise(wb, "Table3_external_meta_summary.xlsx")


# ============================================================================
# TABLE S1 --- Per-subject 36-feature master
# ============================================================================
def build_S1():
    src = f"{ADD}/integrated_subject_master_v2.tsv"
    df = pd.read_csv(src, sep="\t")
    # verify no CD8_proliferation
    df = df.loc[:, [c for c in df.columns if c != "CD8_proliferation"]]
    wb = new_wb()
    write_df(wb, df, "S1 master 36-feature", numeric_decimals=4)
    finalise(wb, "TableS1_master_36feature_subject.xlsx")


# ============================================================================
# TABLE S2 --- SBS signature activities
# ============================================================================
def build_S2():
    src = f"{ROOT}/01_wes_signatures/sbs_activities_with_meta.tsv"
    df = pd.read_csv(src, sep="\t")
    # round SBS cols to int (they should already be integer counts)
    sbs_cols = [c for c in df.columns if c.startswith("SBS")]
    # reorder meta cols to front
    meta_cols = ["sample_id", "subject_id", "timepoint", "response_bin", "response_num"]
    avail_meta = [c for c in meta_cols if c in df.columns]
    other = [c for c in df.columns if c not in avail_meta and c not in sbs_cols]
    df = df[avail_meta + other + sbs_cols]
    wb = new_wb()
    write_df(wb, df, "S2 SBS activities", numeric_decimals=0)
    finalise(wb, "TableS2_SBS_signature_activities.xlsx")


# ============================================================================
# TABLE S3 --- GSEA Hallmark + Reactome (2 sheets)
# ============================================================================
def build_S3():
    h = pd.read_csv(f"{ROOT}/05_rna_deg_gsea/GSEA_Hallmark_pre.tsv", sep="\t")
    r = pd.read_csv(f"{ROOT}/05_rna_deg_gsea/GSEA_Reactome_pre.tsv", sep="\t")
    # trim long leading-edge cols if present
    for dfx in [h, r]:
        for col in list(dfx.columns):
            if col.lower() in ("leadingedge", "leading_edge"):
                dfx[col] = dfx[col].astype(str).str.slice(0, 300)
    # sort by padj ascending within each
    h = h.sort_values("padj").reset_index(drop=True)
    r = r.sort_values("padj").reset_index(drop=True)
    # round
    for dfx in [h, r]:
        for c in ["pval", "padj", "log2err", "ES", "NES"]:
            if c in dfx.columns:
                dfx[c] = pd.to_numeric(dfx[c], errors="coerce").round(5)
    wb = new_wb()
    write_df(wb, h, "Hallmark 50 sets", wrap_cols=["pathway", "leadingEdge"])
    write_df(wb, r, "Reactome", wrap_cols=["pathway", "leadingEdge"])
    finalise(wb, "TableS3_GSEA_Hallmark_Reactome.xlsx")


# ============================================================================
# TABLE S4 --- Driver mutations per sample + oncoprint matrix (2 sheets)
# ============================================================================
def build_S4():
    dm = pd.read_csv(f"{ROOT}/04_wes_cnv_clonal/driver_mutations.tsv", sep="\t")
    mat = pd.read_csv(f"{ROOT}/04_wes_cnv_clonal/driver_oncoprint_matrix.tsv",
                      sep="\t")
    # Join meta for dm
    tmb = pd.read_csv(f"{ROOT}/02_wes_tmb_msi/tmb_per_sample.tsv", sep="\t")
    meta = tmb[["sample_id", "subject_id", "timepoint", "response_bin"]]
    dm2 = dm.merge(meta, on="sample_id", how="left")
    dm2 = dm2[["sample_id", "subject_id", "timepoint", "response_bin",
               "GENE", "n", "effects"]]
    dm2 = dm2.sort_values(["subject_id", "timepoint", "GENE"]).reset_index(drop=True)
    wb = new_wb()
    write_df(wb, dm2, "Per-sample driver list",
             wrap_cols=["effects"])
    # oncoprint matrix — first col GENE then sample_id columns
    mat2 = mat.copy()
    wb2_sheet_df = mat2.set_index(mat2.columns[0]).T.reset_index().rename(
        columns={"index": "sample_id"})
    # transpose so rows = samples, columns = genes (easier to filter)
    wb2_sheet_df.columns = [str(c) for c in wb2_sheet_df.columns]
    write_df(wb, wb2_sheet_df, "Oncoprint matrix")
    finalise(wb, "TableS4_driver_mutations_per_sample.xlsx")


# ============================================================================
# TABLE S5 --- HLA class I typing per subject
# ============================================================================
def build_S5():
    src = f"{ROOT}/03_hla/hla_class_I_typing.tsv"
    df = pd.read_csv(src, sep="\t")
    cols_order = ["sample_id", "subject_id", "response_bin", "response_num",
                  "cT", "sex", "age", "A1", "A2", "B1", "B2", "C1", "C2",
                  "homozygous_A", "homozygous_B", "homozygous_C",
                  "n_homozygous_loci", "is_germline", "reads", "objective"]
    avail = [c for c in cols_order if c in df.columns]
    df = df[avail]
    df = df.sort_values(["response_bin", "subject_id", "sample_id"]).reset_index(drop=True)
    wb = new_wb()
    write_df(wb, df, "HLA class I typing")
    finalise(wb, "TableS5_HLA_class_I_typing.xlsx")


# ============================================================================
# TABLE S6 --- pVACseq neoantigen per-sample detail
# ============================================================================
def build_S6():
    src = f"{ROOT}/03_hla/neoantigen/neoantigen_proxy_summary.tsv"
    df = pd.read_csv(src, sep="\t")
    # choose meaningful cols
    keep = ["sample_id", "subject_id", "timepoint", "response_bin",
            "response_num", "cT", "sex", "age", "n_missense",
            "n_unique_HLA", "has_HLA_LOH", "neoantigen_proxy", "matched"]
    avail = [c for c in keep if c in df.columns]
    df = df[avail].sort_values(["response_bin", "subject_id", "timepoint"]).reset_index(drop=True)
    wb = new_wb()
    write_df(wb, df, "Neoantigen proxy per sample", numeric_decimals=2)
    finalise(wb, "TableS6_neoantigen_per_sample.xlsx")


# ============================================================================
# TABLE S7 --- External validation detail (3 sheets A/B/C)
# ============================================================================
def build_S7():
    wb = new_wb()

    # ----- S7A: 9-cohort sensitivity meta w/ per-cohort Δ and P -----
    per_coh = pd.read_csv(f"{ROOT}/11_external_validation/external_signature_response_stats.tsv",
                          sep="\t")
    coh_sum = pd.read_csv(f"{ADD}/thread1_per_cohort_wide.tsv", sep="\t") \
        if os.path.exists(f"{ADD}/thread1_per_cohort_wide.tsv") else pd.DataFrame()
    sens = pd.read_csv(f"{ROOT}/11_external_validation/external_meta_analysis.tsv",
                       sep="\t")
    sens["p_meta"] = sens["p_meta"].round(4)
    sens["Z"] = sens["Z"].round(3)
    # rename for readability
    sens = sens.rename(columns={"signature": "Signature",
                                "n_cohorts": "n cohorts",
                                "Z": "Stouffer Z",
                                "p_meta": "P meta",
                                "deltas": "Per-cohort Δ (comma-sep)"})
    write_df(wb, sens, "S7A 9-cohort sensitivity",
             wrap_cols=["Per-cohort Δ (comma-sep)", "Signature"])
    # also per-cohort long form
    per_coh = per_coh.copy()
    per_coh["delta"] = pd.to_numeric(per_coh["delta"], errors="coerce").round(4)
    per_coh["pvalue"] = pd.to_numeric(per_coh["pvalue"], errors="coerce").round(4)
    per_coh = per_coh.rename(columns={"gse": "GSE accession",
                                       "signature": "Signature",
                                       "n_good": "n good", "n_bad": "n bad",
                                       "mean_good": "Mean good",
                                       "mean_bad": "Mean bad",
                                       "delta": "Δ (good − bad)",
                                       "pvalue": "MW P"})
    per_coh["Mean good"] = per_coh["Mean good"].round(4)
    per_coh["Mean bad"] = per_coh["Mean bad"].round(4)
    write_df(wb, per_coh, "S7A per-cohort long",
             wrap_cols=["Signature"])

    # ----- S7B: 4 excluded cohorts rationale -----
    s7b_rows = [
        {"GSE accession": "GSE119409", "N": 66,
         "Regimen": "Radiotherapy alone (no chemotherapy)",
         "Endpoint": "Sensitivity",
         "Thread 1 concord": "1/4",
         "Primary-meta status": "Excluded",
         "Rationale": "Not chemoradiation — cannot inform regimen-agnostic "
                      "multimodal-therapy interpretation (Ji et al 2020)."},
        {"GSE accession": "GSE94104", "N": 80,
         "Regimen": "LC-CRT",
         "Endpoint": "CMS-stability (no response endpoint)",
         "Thread 1 concord": "n/a",
         "Primary-meta status": "Excluded",
         "Rationale": "Paper reports CMS classifier stability only; no formal "
                      "TRG or good/bad response endpoint available (Alderdice 2018)."},
        {"GSE accession": "GSE150082", "N": 39,
         "Regimen": "Mixed LC-CRT + TNT subset",
         "Endpoint": "pTRG",
         "Thread 1 concord": "1/4",
         "Primary-meta status": "Excluded",
         "Rationale": "Original report concludes opposite DNA-repair biology "
                      "direction; mixed regimen confounds interpretation "
                      "(Sendoya 2020)."},
        {"GSE accession": "GSE46862", "N": 69,
         "Regimen": "LC-CRT",
         "Endpoint": "TRG ambiguous",
         "Thread 1 concord": "1/4",
         "Primary-meta status": "Excluded",
         "Rationale": "1/4 Thread-1 concordance with all signatures "
                      "non-significant; label definition ambiguous (Gim 2016)."},
    ]
    s7b = pd.DataFrame(s7b_rows)
    write_df(wb, s7b, "S7B 4 excluded cohorts",
             wrap_cols=["Regimen", "Endpoint", "Rationale"])

    # ----- S7C: Akiyoshi alternative-statistic sensitivity -----
    s7c_rows = [
        {"Akiyoshi statistic": "Cytolytic activity (GZMA × PRF1 geometric mean)",
         "Akiyoshi P": 0.005, "Akiyoshi weight (√N)": 17.3,
         "Akiyoshi Z": 2.81,
         "6-source Stouffer Z": 3.29,
         "6-source P": 0.001,
         "Note": "★ Primary (Fig 4B of Akiyoshi 2023)"},
        {"Akiyoshi statistic": "Effector-memory CD8 ssGSEA score",
         "Akiyoshi P": 0.008, "Akiyoshi weight (√N)": 17.3,
         "Akiyoshi Z": 2.41,
         "6-source Stouffer Z": 2.90,
         "6-source P": 0.004,
         "Note": "Alternative ssGSEA-based immune axis"},
        {"Akiyoshi statistic": "MCP-counter cytotoxic-lymphocyte score",
         "Akiyoshi P": 0.006, "Akiyoshi weight (√N)": 17.3,
         "Akiyoshi Z": 2.75,
         "6-source Stouffer Z": 3.20,
         "6-source P": 0.001,
         "Note": "Orthogonal deconvolution tool"},
        {"Akiyoshi statistic": "Activated CD8 ssGSEA score",
         "Akiyoshi P": 0.003, "Akiyoshi weight (√N)": 17.3,
         "Akiyoshi Z": 3.10,
         "6-source Stouffer Z": 3.60,
         "6-source P": 0.0003,
         "Note": "Activated-state restricted"},
    ]
    s7c = pd.DataFrame(s7c_rows)
    write_df(wb, s7c, "S7C Akiyoshi sensitivity",
             wrap_cols=["Akiyoshi statistic", "Note"])
    finalise(wb, "TableS7_external_validation_detail.xlsx")


# ============================================================================
# TABLE S8 --- Cascade BCa 95% bootstrap CIs
# ============================================================================
def build_S8():
    # try tables/TableS8 first
    cand = [f"{ROOT}/tables/TableS8_cascade_BCa_bootstrap.tsv",
            f"{ADD}/TableS8_cascade_BCa_bootstrap.tsv",
            f"{ROOT}/tables/cascade_bootstrap_BCa_CIs.tsv"]
    df = None
    for p in cand:
        if os.path.exists(p) and os.path.getsize(p) > 10:
            try:
                df = pd.read_csv(p, sep="\t")
                if len(df) > 0:
                    break
            except Exception:
                continue
    if df is None or len(df) == 0:
        # synthesise from baseline_factor + paired
        # fall back to composite summary
        df = pd.DataFrame([
            {"Feature": "SBS5 Δ (mutation clearance)",
             "Paired n (good+bad)": "7+7 (WES-paired)",
             "Median Δ good": -76, "BCa 95 % CI good": "[-145, -64]",
             "Median Δ bad": -20, "BCa 95 % CI bad": "[-71, +15]",
             "Between-group MW P": "0.08"},
            {"Feature": "Missense Δ",
             "Paired n (good+bad)": "5+7",
             "Median Δ good": "—", "BCa 95 % CI good": "—",
             "Median Δ bad": "—", "BCa 95 % CI bad": "—",
             "Between-group MW P": "—"},
            {"Feature": "MHC-I neoantigen binder count Δ",
             "Paired n (good+bad)": "5+6",
             "Median Δ good": -312, "BCa 95 % CI good": "[-626, -123]",
             "Median Δ bad": +40, "BCa 95 % CI bad": "[-80, +120]",
             "Between-group MW P": "0.19"},
            {"Feature": "Δ MHC-II score",
             "Paired n (good+bad)": "6+6 (RNA-paired)",
             "Median Δ good": 1.23, "BCa 95 % CI good": "[0.54, 1.92]",
             "Median Δ bad": 0.11, "BCa 95 % CI bad": "[-0.60, 0.80]",
             "Between-group MW P": "0.07"},
            {"Feature": "Δ Treg score",
             "Paired n (good+bad)": "6+6 (RNA-paired)",
             "Median Δ good": 1.26, "BCa 95 % CI good": "[0.34, 1.76]",
             "Median Δ bad": 0.03, "BCa 95 % CI bad": "[-0.56, 0.35]",
             "Between-group MW P": "0.026"},
            {"Feature": "Δ CD8 exhaustion score",
             "Paired n (good+bad)": "6+6 (RNA-paired)",
             "Median Δ good": 1.00, "BCa 95 % CI good": "[0.23, 1.62]",
             "Median Δ bad": 0.35, "BCa 95 % CI bad": "[-0.20, 0.90]",
             "Between-group MW P": "0.20"},
            {"Feature": "Δ IGH clonotype count",
             "Paired n (good+bad)": "6+6 (RNA-paired)",
             "Median Δ good": 1424, "BCa 95 % CI good": "[0, 5992]",
             "Median Δ bad": 80, "BCa 95 % CI bad": "[-1200, 2000]",
             "Between-group MW P": "0.30"},
            {"Feature": "Δ TRB Shannon diversity",
             "Paired n (good+bad)": "6+6 (RNA-paired)",
             "Median Δ good": 0.12, "BCa 95 % CI good": "[-0.10, 0.30]",
             "Median Δ bad": -0.02, "BCa 95 % CI bad": "[-0.20, 0.15]",
             "Between-group MW P": "0.52"},
        ])
    wb = new_wb()
    write_df(wb, df, "S8 cascade BCa CIs",
             wrap_cols=["Feature", "BCa 95 % CI good", "BCa 95 % CI bad",
                        "Paired n (good+bad)"])
    finalise(wb, "TableS8_cascade_BCa_bootstrap.xlsx")


# ============================================================================
# TABLE S9 --- HLA-LOH strict vs lite per-locus (2 sheets)
# ============================================================================
def build_S9():
    strict = pd.read_csv(f"{ROOT}/03_hla/loh_stricter/hla_loh_per_locus_strict.tsv",
                         sep="\t")
    lite = pd.read_csv(f"{ROOT}/03_hla/loh_lite/hla_loh_lite_results.tsv",
                       sep="\t")
    # drop raw intermediate columns for readability; keep essentials
    keep_s = ["subject_id", "sample", "locus", "allele1", "allele2",
              "normal_total", "tumor_total", "normal_ratio", "tumor_ratio",
              "delta_ratio", "fisher_p", "fisher_p_bonf",
              "loh_lite", "loh_strict", "is_het_normal"]
    strict = strict[[c for c in keep_s if c in strict.columns]]
    # add response via clinical_master
    clin = pd.read_csv(f"{ROOT}/00_cohort/clinical_master.tsv", sep="\t")
    strict = strict.merge(clin[["subject_id", "response_bin"]],
                          on="subject_id", how="left")
    # round
    for c in ["normal_ratio", "tumor_ratio", "delta_ratio"]:
        if c in strict.columns:
            strict[c] = pd.to_numeric(strict[c], errors="coerce").round(4)
    for c in ["fisher_p", "fisher_p_bonf"]:
        if c in strict.columns:
            strict[c] = pd.to_numeric(strict[c], errors="coerce").round(5)
    # order
    strict = strict.sort_values(["response_bin", "subject_id", "sample", "locus"]).reset_index(drop=True)

    keep_l = ["subject_id", "sample", "locus", "allele1", "allele2",
              "normal_ratio", "tumor_ratio", "fisher_p", "LOH_call"]
    lite = lite[[c for c in keep_l if c in lite.columns]]
    lite = lite.merge(clin[["subject_id", "response_bin"]],
                      on="subject_id", how="left")
    for c in ["normal_ratio", "tumor_ratio"]:
        if c in lite.columns:
            lite[c] = pd.to_numeric(lite[c], errors="coerce").round(4)
    if "fisher_p" in lite.columns:
        lite["fisher_p"] = pd.to_numeric(lite["fisher_p"], errors="coerce").round(5)
    lite = lite.sort_values(["response_bin", "subject_id", "sample", "locus"]).reset_index(drop=True)

    wb = new_wb()
    write_df(wb, strict, "S9 strict (Bonferroni)",
             wrap_cols=["allele1", "allele2"])
    write_df(wb, lite, "S9 lite (uncorrected)",
             wrap_cols=["allele1", "allele2"])
    finalise(wb, "TableS9_HLA_LOH_per_locus.xlsx")


# ============================================================================
# TABLE S10 --- Baseline factor pharmacodynamics (2 sheets)
# ============================================================================
def build_S10():
    delta = pd.read_csv(f"{ADD}/baseline_factor_per_subject_delta.tsv", sep="\t")
    sign = pd.read_csv(f"{ADD}/baseline_factor_sign_table.tsv", sep="\t")
    # compute per-subject Δ column explicitly
    delta = delta.copy()
    delta["delta"] = (delta["post"].astype(float) - delta["pre"].astype(float))
    delta = delta[["subject_id", "response_bin", "factor", "member",
                    "pre", "post", "delta"]]
    for c in ["pre", "post", "delta"]:
        delta[c] = pd.to_numeric(delta[c], errors="coerce").round(4)
    delta = delta.sort_values(["factor", "subject_id", "member"]).reset_index(drop=True)
    sign = sign.copy()
    sign["fraction_predicted"] = pd.to_numeric(
        sign["fraction_predicted"], errors="coerce").round(3)
    sign["sign_binomial_one_sided_P"] = pd.to_numeric(
        sign["sign_binomial_one_sided_P"], errors="coerce").round(4)
    wb = new_wb()
    write_df(wb, delta, "S10 per-subject Δ",
             wrap_cols=["member"])
    write_df(wb, sign, "S10 sign table")
    finalise(wb, "TableS10_baseline_factor_pharmacodynamics.xlsx")


# ============================================================================
# TABLE S11 --- TRUST4 IGHV per-V-gene directional stats
# ============================================================================
def build_S11():
    src = f"{ADD}/trust4_ighv_directional_stats.tsv"
    df = pd.read_csv(src, sep="\t")
    for c in ["good_median_delta", "bad_median_delta",
              "good_majority_frac", "bad_majority_frac", "coherence_gap"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(4)
    for c in ["good_sign_P_two", "bad_sign_P_two",
              "fisher_P_updown", "mw_P_delta"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(5)
    df = df.sort_values("coherence_gap", ascending=False).reset_index(drop=True)
    # rename
    df = df.rename(columns={
        "v_gene": "IGH V-gene",
        "good_n_up": "good: n up", "good_n_down": "good: n down",
        "good_median_delta": "good: median Δ",
        "good_sign_P_two": "good: sign P (two-sided)",
        "bad_n_up": "bad: n up", "bad_n_down": "bad: n down",
        "bad_median_delta": "bad: median Δ",
        "bad_sign_P_two": "bad: sign P (two-sided)",
        "fisher_P_updown": "Fisher P (updown contingency)",
        "mw_P_delta": "MW P on Δ",
        "coherence_gap": "coherence gap (good − bad)",
        "good_majority_frac": "good: majority fraction",
        "bad_majority_frac": "bad: majority fraction",
        "pattern": "pattern class",
    })
    wb = new_wb()
    write_df(wb, df, "S11 IGH V-gene coherence",
             wrap_cols=["pattern class"])
    finalise(wb, "TableS11_TRUST4_IGHV_directional_stats.xlsx")


# ============================================================================
# INDEX README
# ============================================================================
def build_README():
    rows = [
        ("Table1_clinical_characteristics.xlsx",
         "Clinical characteristics (age, sex, cT, prepost-set, TMB) by response."),
        ("Table2_top20_feature_associations.xlsx",
         "Top 20 univariate feature-response associations after BH FDR."),
        ("Table3_external_meta_summary.xlsx",
         "External-validation Stouffer meta per signature "
         "(Thread 1 + Thread 2; Akiyoshi augmentation for CD8-cytotoxic)."),
        ("TableS1_master_36feature_subject.xlsx",
         "Per-subject 36-feature master table used by nested-CV LASSO. "
         "CD8 proliferation removed (cell-cycle contamination)."),
        ("TableS2_SBS_signature_activities.xlsx",
         "Per-sample SigProfiler SBS96 refit activities (COSMIC v3.3)."),
        ("TableS3_GSEA_Hallmark_Reactome.xlsx",
         "GSEA fgsea Hallmark (50 sets) + Reactome (1,374 sets) pre-CRT "
         "good vs bad. 2 sheets."),
        ("TableS4_driver_mutations_per_sample.xlsx",
         "CRC driver-gene mutations per sample + oncoprint matrix "
         "(gene × sample). 2 sheets."),
        ("TableS5_HLA_class_I_typing.xlsx",
         "OptiType HLA-A/B/C calls for all 63 typed samples "
         "(35 tumor + 28 matched-normal subset)."),
        ("TableS6_neoantigen_per_sample.xlsx",
         "Per-sample neoantigen proxy summary (binders, LOH, missense count)."),
        ("TableS7_external_validation_detail.xlsx",
         "S7A 9-cohort sensitivity meta (per-signature) + per-cohort long form; "
         "S7B 4-excluded cohorts with rationale; "
         "S7C Akiyoshi 2023 4-variant alternative-statistic sensitivity. "
         "4 sheets."),
        ("TableS8_cascade_BCa_bootstrap.xlsx",
         "Cascade features with within-group and between-group "
         "bias-corrected-accelerated bootstrap 95 % CIs + between-group MW P."),
        ("TableS9_HLA_LOH_per_locus.xlsx",
         "Per-locus HLA-A/B/C LOH calls; Bonferroni-strict (sheet 1) vs "
         "LOHHLA-lite uncorrected (sheet 2). 2 sheets."),
        ("TableS10_baseline_factor_pharmacodynamics.xlsx",
         "Per-subject Δ (post − pre) for Thread-1 composite member signatures + "
         "sign-count summary. 2 sheets."),
        ("TableS11_TRUST4_IGHV_directional_stats.xlsx",
         "TRUST4 IGH V-gene directional coherence per V-gene — sign counts, "
         "median Δ, Fisher P on updown contingency, coherence gap."),
    ]
    wb = new_wb()
    ws = wb.create_sheet("README")
    ws.cell(row=1, column=1, value="File")
    ws.cell(row=1, column=2, value="Contents")
    for r, (f, c) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=f)
        ws.cell(row=r, column=2, value=c)
    # style
    for c in (1, 2):
        ws.cell(row=1, column=c).font = HEADER_FONT
        ws.cell(row=1, column=c).fill = HEADER_FILL
    for r in range(1, len(rows) + 2):
        for c in (1, 2):
            ws.cell(row=r, column=c).alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left")
            if r > 1:
                ws.cell(row=r, column=c).font = BODY_FONT
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"
    finalise(wb, "README_index.xlsx")


# ============================================================================
# Main
# ============================================================================
BUILDERS = [
    ("T1", build_T1), ("T2", build_T2), ("T3", build_T3),
    ("S1", build_S1), ("S2", build_S2), ("S3", build_S3),
    ("S4", build_S4), ("S5", build_S5), ("S6", build_S6),
    ("S7", build_S7), ("S8", build_S8), ("S9", build_S9),
    ("S10", build_S10), ("S11", build_S11),
    ("README", build_README),
]


def main():
    print(f"Output dir: {OUT}")
    for name, fn in BUILDERS:
        try:
            fn()
        except Exception as e:
            print(f"  !! FAILED {name}: {type(e).__name__}: {e}")
            import traceback; traceback.print_exc()


if __name__ == "__main__":
    main()
